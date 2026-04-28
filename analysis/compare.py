"""Builds Excel comparison report from extracted SLI data."""

import pandas as pd
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path(__file__).parent.parent / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

WAVE_ORDER = ["March 2025", "September 2025", "March 2026"]
HEADER_COLOR = "1F4E79"
ALT_COLOR = "D6E4F0"


def _style_ws(ws):
    fill = PatternFill("solid", fgColor=HEADER_COLOR)
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


def build_excel(rows: list[dict], path: Path = None) -> Path:
    if not rows:
        print("No data to export.")
        return None

    df = pd.DataFrame(rows)
    df["wave_label"] = pd.Categorical(df["wave_label"], categories=WAVE_ORDER, ordered=True)

    # Only numeric rows for pivot tables
    df_num = df[df["value"].notna()].copy()
    df_num["value"] = pd.to_numeric(df_num["value"], errors="coerce")

    out = path or OUTPUT_DIR / "disinfocode_comparativa.xlsx"

    with pd.ExcelWriter(out, engine="openpyxl") as writer:

        # ── 1. Datos completos (todas las SLIs con tabla) ─────────────────────
        export_cols = ["wave_label", "platform", "service", "chapter",
                       "commitment", "measure", "sli_name", "country",
                       "metric_name", "value", "raw_value", "methodology"]
        rename = {
            "wave_label": "Ola", "platform": "Plataforma", "service": "Servicio",
            "chapter": "Capítulo", "commitment": "Compromiso", "measure": "Medida",
            "sli_name": "SLI", "country": "País", "metric_name": "Métrica",
            "value": "Valor numérico", "raw_value": "Valor original",
            "methodology": "Metodología",
        }
        df[export_cols].rename(columns=rename).to_excel(
            writer, sheet_name="Datos completos", index=False)
        _style_ws(writer.sheets["Datos completos"])

        # ── 2. Resumen por Plataforma × Ola ───────────────────────────────────
        if not df_num.empty:
            piv = df_num.pivot_table(
                index=["chapter", "sli_name", "metric_name"],
                columns=["wave_label", "platform"],
                values="value",
                aggfunc="sum",
            )
            piv.to_excel(writer, sheet_name="Por Plataforma × Ola")
            _style_ws(writer.sheets["Por Plataforma × Ola"])

        # ── 3. Comparativa por País (totales EU/EEA) ──────────────────────────
        totals = df_num[df_num["country"].str.contains("Total", na=False)].copy()
        if not totals.empty:
            piv_country = totals.pivot_table(
                index=["sli_name", "metric_name", "country"],
                columns=["wave_label", "platform"],
                values="value",
                aggfunc="sum",
            )
            piv_country.to_excel(writer, sheet_name="Totales EU-EEA")
            _style_ws(writer.sheets["Totales EU-EEA"])

        # ── 4. Por País (desglose por estado miembro) ─────────────────────────
        eu_rows = df_num[~df_num["country"].str.contains("Total", na=False)].copy()
        if not eu_rows.empty:
            piv_eu = eu_rows.pivot_table(
                index=["country"],
                columns=["wave_label", "platform"],
                values="value",
                aggfunc="sum",
            )
            piv_eu.to_excel(writer, sheet_name="Por País")
            _style_ws(writer.sheets["Por País"])

        # ── 5. Por SLI individual: una hoja por código SLI ───────────────────
        sli_codes = df_num["sli_code"].dropna().unique()
        for code in sorted(sli_codes):
            df_sli = df_num[df_num["sli_code"] == code]
            sli_label = df_sli["sli_name"].iloc[0] if not df_sli.empty else code
            sheet_name = f"SLI {code}"[:31]
            try:
                piv_sli = df_sli.pivot_table(
                    index=["country"],
                    columns=["wave_label", "platform"],
                    values="value",
                    aggfunc="sum",
                ).reindex(columns=pd.MultiIndex.from_product(
                    [WAVE_ORDER, df_sli["platform"].unique()],
                    names=["wave_label", "platform"]
                ), fill_value=None)
                piv_sli.to_excel(writer, sheet_name=sheet_name)
                ws = writer.sheets[sheet_name]
                _style_ws(ws)
                # Add SLI description in first row
                ws.insert_rows(1)
                ws["A1"] = sli_label
                ws["A1"].font = Font(bold=True, color="1F4E79")
            except Exception:
                pass

        # ── 6. CSV plano para análisis externo ────────────────────────────────
        df_num[["wave_label", "platform", "service", "sli_code", "sli_name",
                "chapter", "commitment_code", "measure_code",
                "metric_name", "country", "value"]].to_excel(
            writer, sheet_name="Para análisis R-Python", index=False)
        _style_ws(writer.sheets["Para análisis R-Python"])

    print(f"\nExcel generado: {out}")
    return out


def save_csv(rows: list[dict], path: Path = None) -> Path:
    if not rows:
        return None
    df = pd.DataFrame(rows)
    out = path or OUTPUT_DIR / "metrics_raw.csv"
    df.to_csv(out, index=False, encoding="utf-8-sig")
    print(f"CSV guardado: {out}")
    return out
